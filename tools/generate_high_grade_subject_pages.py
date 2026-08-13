from __future__ import annotations

import argparse
import hashlib
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook

import generate_korean_english_math_pages as shared
import generate_sitemap_robots
import high_grade_keyword_strategy as keyword_strategy


SOURCE_DIR = Path.home() / "Desktop" / "새 폴더"
EXISTING_PUBLISHED_AT = "2026-08-13T00:00:00+09:00"
ELEMENTARY_PUBLISHED_AT = "2026-08-14T00:00:00+09:00"
UPDATED_AT = "2026-08-14T00:00:00+09:00"


@dataclass(frozen=True)
class CategoryConfig:
    category: str
    grade: str
    subject: str
    source_name: str
    card_label: str
    card_description: str
    detail_eyebrow: str
    category_eyebrow: str
    category_description: str
    aside_title: str
    transition: str
    parent_title: str
    parent_copy: str
    school_level: str


CONFIGS = (
    CategoryConfig(
        category="고1수학학원",
        grade="고1",
        subject="수학",
        source_name="고1 수학학원 원고.xlsx",
        card_label="고1 · 수학 내신 · 고교 전환",
        card_description="전국 371개 지역별 고1 수학 진단·내신 학습 안내",
        detail_eyebrow="GRADE 10 MATH LOCAL GUIDE",
        category_eyebrow="GRADE 10 MATH DIRECTORY",
        category_description=(
            "전국 371개 동네별 고1수학학원 선택 기준을 정리했습니다. "
            "중학교 수학에서 고교 수학으로 넘어갈 때 필요한 개념 연결, 조건 해석, "
            "학교 시험과 오답 재학습 기준을 지역별로 확인할 수 있습니다."
        ),
        aside_title="고교 수학의 첫 단원보다 현재 풀이에서 끊기는 지점을 먼저 찾습니다.",
        transition="중학교식 풀이에서 고교 내신형 설명으로 넘어가는 시기",
        parent_title="고1 수학의 개념 연결과 내신 준비",
        parent_copy="최근 시험지와 풀이 설명을 바탕으로 개념·조건 해석·오답 재풀이 순서를 정리했습니다.",
        school_level="고등",
    ),
    CategoryConfig(
        category="고1영어학원",
        grade="고1",
        subject="영어",
        source_name="고1 영어학원 원고.xlsx",
        card_label="고1 · 공통영어 · 내신 전환",
        card_description="전국 371개 지역별 고1 영어 진단·내신 학습 안내",
        detail_eyebrow="GRADE 10 ENGLISH LOCAL GUIDE",
        category_eyebrow="GRADE 10 ENGLISH DIRECTORY",
        category_description=(
            "전국 371개 동네별 고1영어학원 선택 기준을 정리했습니다. "
            "고교 영어의 어휘 누적, 문장 구조, 학교 본문과 독해 근거, 서술형 준비를 "
            "지역·학교 일정과 함께 확인할 수 있습니다."
        ),
        aside_title="암기량보다 문장 구조와 답의 근거를 학생이 설명할 수 있는지 먼저 봅니다.",
        transition="중학교 영어에서 고교 내신형 독해와 서술형으로 넘어가는 시기",
        parent_title="고1 영어의 문장 구조와 내신 준비",
        parent_copy="어휘·문법·독해를 따로 진단하고 학교 본문과 서술형 준비 순서로 연결했습니다.",
        school_level="고등",
    ),
    CategoryConfig(
        category="고2수학학원",
        grade="고2",
        subject="수학",
        source_name="고2 수학학원 원고.xlsx",
        card_label="고2 · 수학 내신 · 누적 개념",
        card_description="전국 371개 지역별 고2 수학 진단·내신 학습 안내",
        detail_eyebrow="GRADE 11 MATH LOCAL GUIDE",
        category_eyebrow="GRADE 11 MATH DIRECTORY",
        category_description=(
            "전국 371개 동네별 고2수학학원 선택 기준을 정리했습니다. "
            "학교별 진도와 누적 개념, 조건 해석, 풀이 과정, 시험 뒤 오답 재확인을 "
            "학생의 최근 자료에 맞춰 살펴볼 수 있습니다."
        ),
        aside_title="새 단원 진도와 누적 빈틈을 구분해 고2 수학의 우선순위를 정합니다.",
        transition="학교별 진도가 빨라지고 이전 단원의 빈틈이 함께 드러나는 시기",
        parent_title="고2 수학의 누적 개념과 시험 준비",
        parent_copy="현재 단원과 이전 단원의 빈틈을 나누고 풀이 기록을 다음 시험 계획으로 연결했습니다.",
        school_level="고등",
    ),
    CategoryConfig(
        category="고2영어학원",
        grade="고2",
        subject="영어",
        source_name="고2 영어학원 원고.xlsx",
        card_label="고2 · 영어 내신 · 독해 근거",
        card_description="전국 371개 지역별 고2 영어 진단·내신 학습 안내",
        detail_eyebrow="GRADE 11 ENGLISH LOCAL GUIDE",
        category_eyebrow="GRADE 11 ENGLISH DIRECTORY",
        category_description=(
            "전국 371개 동네별 고2영어학원 선택 기준을 정리했습니다. "
            "학교 본문과 어휘 누적, 문장 구조, 독해 근거, 서술형 교정을 최근 시험 "
            "자료와 연결해 확인할 수 있습니다."
        ),
        aside_title="점수보다 어휘·문장 구조·독해 근거 중 어디에서 막혔는지 먼저 나눕니다.",
        transition="내신 지문과 독해 난도가 높아지고 누적 어휘의 차이가 커지는 시기",
        parent_title="고2 영어의 독해 근거와 내신 준비",
        parent_copy="학교 본문·어휘·문장 구조·서술형 기록을 나누어 다음 시험의 보완 순서를 정리했습니다.",
        school_level="고등",
    ),
    CategoryConfig(
        category="중1수학학원",
        grade="중1",
        subject="수학",
        source_name="중1 수학학원 원고.xlsx",
        card_label="중1 · 수학 내신 · 중등 전환",
        card_description="전국 371개 지역별 중1 수학 진단·내신 학습 안내",
        detail_eyebrow="GRADE 7 MATH LOCAL GUIDE",
        category_eyebrow="GRADE 7 MATH DIRECTORY",
        category_description=(
            "전국 371개 동네별 중1수학학원 선택 기준을 정리했습니다. "
            "초등 수학에서 중등 수학으로 넘어갈 때 필요한 개념어, 문자와 식, "
            "학교 시험과 오답 재학습 기준을 지역별로 확인할 수 있습니다."
        ),
        aside_title="문제 수보다 초등 과정의 빈틈과 중1 개념어를 연결하는 힘을 먼저 봅니다.",
        transition="초등식 계산에서 중학교 개념과 풀이 설명으로 넘어가는 시기",
        parent_title="중1 수학의 개념 전환과 첫 내신 준비",
        parent_copy="초등 과정의 빈틈과 중1 개념어·문자식·풀이 기록을 연결해 첫 시험 준비 순서를 정리했습니다.",
        school_level="중등",
    ),
    CategoryConfig(
        category="중1영어학원",
        grade="중1",
        subject="영어",
        source_name="중1 영어학원 원고.xlsx",
        card_label="중1 · 영어 내신 · 문장 전환",
        card_description="전국 371개 지역별 중1 영어 진단·내신 학습 안내",
        detail_eyebrow="GRADE 7 ENGLISH LOCAL GUIDE",
        category_eyebrow="GRADE 7 ENGLISH DIRECTORY",
        category_description=(
            "전국 371개 동네별 중1영어학원 선택 기준을 정리했습니다. "
            "초등 영어에서 중학교 어휘·문법·본문 학습으로 넘어갈 때 필요한 진단과 "
            "학교 내신·서술형 준비 기준을 확인할 수 있습니다."
        ),
        aside_title="단어 암기량보다 중학교 문장의 구조와 답의 근거를 설명하는지 먼저 봅니다.",
        transition="초등 영어에서 중학교 어휘·문법·본문 평가로 넘어가는 시기",
        parent_title="중1 영어의 문장 구조와 첫 내신 준비",
        parent_copy="어휘·문법·교과서 본문을 따로 진단하고 중학교 첫 시험과 서술형 준비로 연결했습니다.",
        school_level="중등",
    ),
    CategoryConfig(
        category="중2수학학원",
        grade="중2",
        subject="수학",
        source_name="중2 수학학원 원고.xlsx",
        card_label="중2 · 수학 내신 · 누적 개념",
        card_description="전국 371개 지역별 중2 수학 진단·내신 학습 안내",
        detail_eyebrow="GRADE 8 MATH LOCAL GUIDE",
        category_eyebrow="GRADE 8 MATH DIRECTORY",
        category_description=(
            "전국 371개 동네별 중2수학학원 선택 기준을 정리했습니다. "
            "중1 과정의 누적 빈틈과 중2 개념, 조건 해석, 풀이 과정, 시험 뒤 "
            "오답 재확인 기준을 최근 자료에 맞춰 살펴볼 수 있습니다."
        ),
        aside_title="현재 단원과 중1 과정에서 남은 빈틈을 구분해 중2 수학의 우선순위를 정합니다.",
        transition="문자식·함수·도형의 연결이 늘고 누적 개념의 차이가 드러나는 시기",
        parent_title="중2 수학의 누적 개념과 내신 준비",
        parent_copy="현재 단원과 이전 과정의 빈틈을 나누고 풀이 기록을 다음 시험 계획으로 연결했습니다.",
        school_level="중등",
    ),
    CategoryConfig(
        category="중2영어학원",
        grade="중2",
        subject="영어",
        source_name="중2 영어학원 원고.xlsx",
        card_label="중2 · 영어 내신 · 독해 근거",
        card_description="전국 371개 지역별 중2 영어 진단·내신 학습 안내",
        detail_eyebrow="GRADE 8 ENGLISH LOCAL GUIDE",
        category_eyebrow="GRADE 8 ENGLISH DIRECTORY",
        category_description=(
            "전국 371개 동네별 중2영어학원 선택 기준을 정리했습니다. "
            "누적 어휘와 문장 구조, 교과서 본문, 독해 근거와 서술형 교정을 "
            "학교 시험 자료와 연결해 확인할 수 있습니다."
        ),
        aside_title="점수보다 어휘·문장 구조·본문 근거 중 어디에서 막혔는지 먼저 나눕니다.",
        transition="본문과 문법의 난도가 높아지고 서술형 설명의 차이가 커지는 시기",
        parent_title="중2 영어의 독해 근거와 내신 준비",
        parent_copy="어휘·문장 구조·학교 본문·서술형 기록을 나누어 다음 시험의 보완 순서를 정리했습니다.",
        school_level="중등",
    ),
    CategoryConfig(
        category="중3수학학원",
        grade="중3",
        subject="수학",
        source_name="중3 수학학원 원고.xlsx",
        card_label="중3 · 수학 내신 · 고교 준비",
        card_description="전국 371개 지역별 중3 수학 진단·내신 학습 안내",
        detail_eyebrow="GRADE 9 MATH LOCAL GUIDE",
        category_eyebrow="GRADE 9 MATH DIRECTORY",
        category_description=(
            "전국 371개 동네별 중3수학학원 선택 기준을 정리했습니다. "
            "중학교 마지막 내신과 누적 개념, 고교 수학 전환에 필요한 조건 해석, "
            "풀이 설명과 오답 재학습 기준을 확인할 수 있습니다."
        ),
        aside_title="중3 현재 단원과 고교 수학으로 이어질 누적 개념을 구분해 우선순위를 정합니다.",
        transition="중학교 내신을 마무리하며 고교 수학의 개념 언어를 준비하는 시기",
        parent_title="중3 수학의 내신 마무리와 고교 준비",
        parent_copy="중학교 누적 빈틈과 현재 단원을 나누고 풀이 설명을 고교 전환 준비로 연결했습니다.",
        school_level="중등",
    ),
    CategoryConfig(
        category="중3영어학원",
        grade="중3",
        subject="영어",
        source_name="중3 영어학원 원고.xlsx",
        card_label="중3 · 영어 내신 · 고교 전환",
        card_description="전국 371개 지역별 중3 영어 진단·내신 학습 안내",
        detail_eyebrow="GRADE 9 ENGLISH LOCAL GUIDE",
        category_eyebrow="GRADE 9 ENGLISH DIRECTORY",
        category_description=(
            "전국 371개 동네별 중3영어학원 선택 기준을 정리했습니다. "
            "중학교 내신을 마무리하면서 누적 어휘, 문장 구조, 독해 근거, 서술형 "
            "기록을 고교 영어 준비와 연결해 확인할 수 있습니다."
        ),
        aside_title="중학교 점수보다 고교 지문으로 이어질 어휘·구조·근거의 빈틈을 먼저 봅니다.",
        transition="중학교 내신을 마무리하며 고교 독해와 서술형을 준비하는 시기",
        parent_title="중3 영어의 내신 마무리와 고교 전환",
        parent_copy="어휘·문장 구조·독해 근거·서술형 기록을 나누어 고교 영어의 시작 순서를 정리했습니다.",
        school_level="중등",
    ),
    CategoryConfig(
        category="초3수학학원",
        grade="초3",
        subject="수학",
        source_name="초3 수학학원 원고.xlsx",
        card_label="초3 · 수학 기초 · 연산 원리",
        card_description="전국 371개 지역별 초3 수학 진단·복습 안내",
        detail_eyebrow="GRADE 3 MATH LOCAL GUIDE",
        category_eyebrow="GRADE 3 MATH DIRECTORY",
        category_description=(
            "전국 371개 동네별 초3수학학원 선택 기준을 정리했습니다. "
            "곱셈·나눗셈과 분수·도형을 외우기보다 원리를 설명하는지, 문장제의 "
            "조건을 읽고 풀이를 이어 가는지를 학교 자료와 함께 확인할 수 있습니다."
        ),
        aside_title="정답 수보다 연산 원리와 문제의 조건을 학생이 설명하는지 먼저 봅니다.",
        transition="기초 연산을 바탕으로 곱셈·나눗셈과 분수·도형의 원리를 넓혀 가는 시기",
        parent_title="초3 수학의 연산 원리와 문제 읽기",
        parent_copy="연산 과정·문장제 조건·단원평가 기록을 나누어 첫 보완 순서를 정리했습니다.",
        school_level="초등",
    ),
    CategoryConfig(
        category="초3영어학원",
        grade="초3",
        subject="영어",
        source_name="초3 영어학원.xlsx",
        card_label="초3 · 영어 시작 · 소리와 문장",
        card_description="전국 371개 지역별 초3 영어 진단·복습 안내",
        detail_eyebrow="GRADE 3 ENGLISH LOCAL GUIDE",
        category_eyebrow="GRADE 3 ENGLISH DIRECTORY",
        category_description=(
            "전국 371개 동네별 초3영어학원 선택 기준을 정리했습니다. "
            "소리와 철자, 기본 어휘와 짧은 문장을 듣고 말하고 읽는 과정을 "
            "학교 진도와 가정 복습 기록에 맞춰 확인할 수 있습니다."
        ),
        aside_title="단어 수보다 소리·철자·뜻을 연결하고 짧은 문장을 이해하는지 먼저 봅니다.",
        transition="학교 영어를 시작하며 소리·철자·뜻과 기본 문장 표현을 연결하는 시기",
        parent_title="초3 영어의 소리·철자 연결과 첫 문장",
        parent_copy="듣기·말하기·읽기 기록을 나누고 짧은 가정 복습으로 이어지는 순서를 정리했습니다.",
        school_level="초등",
    ),
    CategoryConfig(
        category="초4수학학원",
        grade="초4",
        subject="수학",
        source_name="초4 수학학원 원고.xlsx",
        card_label="초4 · 수학 응용 · 분수와 도형",
        card_description="전국 371개 지역별 초4 수학 진단·복습 안내",
        detail_eyebrow="GRADE 4 MATH LOCAL GUIDE",
        category_eyebrow="GRADE 4 MATH DIRECTORY",
        category_description=(
            "전국 371개 동네별 초4수학학원 선택 기준을 정리했습니다. "
            "여러 단계의 계산과 문장제, 분수·소수와 도형 개념을 학생이 설명하고 "
            "다시 적용하는지를 최근 학습 자료로 확인할 수 있습니다."
        ),
        aside_title="풀이 속도보다 여러 단계의 조건을 정리하고 계산 이유를 설명하는지 먼저 봅니다.",
        transition="연산 정확도와 개념 설명을 바탕으로 여러 단계의 문제에 적용하는 힘을 키우는 시기",
        parent_title="초4 수학의 개념 적용과 문장제 풀이",
        parent_copy="분수·소수·도형과 문장제의 막힌 단계를 나누고 재풀이 순서를 정리했습니다.",
        school_level="초등",
    ),
    CategoryConfig(
        category="초4영어학원",
        grade="초4",
        subject="영어",
        source_name="초4 영어학원 원고.xlsx",
        card_label="초4 · 영어 문장 · 읽기 습관",
        card_description="전국 371개 지역별 초4 영어 진단·복습 안내",
        detail_eyebrow="GRADE 4 ENGLISH LOCAL GUIDE",
        category_eyebrow="GRADE 4 ENGLISH DIRECTORY",
        category_description=(
            "전국 371개 동네별 초4영어학원 선택 기준을 정리했습니다. "
            "누적 어휘와 기본 문장 구조, 듣기·말하기·읽기·쓰기의 연결 상태를 "
            "학교 자료와 주간 복습 기록으로 살펴볼 수 있습니다."
        ),
        aside_title="암기한 단어보다 문장 속 뜻을 이해하고 말과 글로 다시 쓰는지 먼저 봅니다.",
        transition="익힌 낱말과 표현을 짧은 문장 읽기·말하기·쓰기로 넓혀 가는 시기",
        parent_title="초4 영어의 기본 문장과 읽기 습관",
        parent_copy="어휘·문장 구조·읽기와 쓰기 기록을 나누어 다음 복습 순서를 정리했습니다.",
        school_level="초등",
    ),
)


SIGNALS = {
    "수학": (
        {
            "label": "개념 연결",
            "keywords": ("개념", "공식", "정의"),
            "persona": "공식은 기억하지만 개념 사이의 연결을 설명하기 어려운",
            "evidence": "개념 문항의 풀이 설명과 교재 진도",
            "action": "개념을 말로 설명한 뒤 대표 문항을 다시 푸는 순서",
        },
        {
            "label": "조건 해석",
            "keywords": ("조건", "해석", "식 세우"),
            "persona": "문제의 조건을 식으로 옮기는 단계에서 자주 멈추는",
            "evidence": "조건 표시와 식을 세운 흔적",
            "action": "조건을 표시하고 식의 근거를 한 줄씩 확인하는 순서",
        },
        {
            "label": "계산 정확도",
            "keywords": ("계산", "실수", "부호"),
            "persona": "풀이 방향은 맞아도 계산과 부호 실수가 반복되는",
            "evidence": "계산 과정과 틀린 지점을 고친 기록",
            "action": "계산 단계를 줄 단위로 나누고 검산 기준을 남기는 순서",
        },
        {
            "label": "함수·그래프",
            "keywords": ("함수", "그래프", "좌표"),
            "persona": "식과 그래프의 관계를 문제마다 새로 외우려는",
            "evidence": "함수식과 그래프를 연결해 설명한 흔적",
            "action": "식·표·그래프를 같은 조건으로 바꾸어 보는 순서",
        },
        {
            "label": "서술형 풀이",
            "keywords": ("서술", "풀이 과정", "과정"),
            "persona": "답은 맞혀도 풀이 근거를 문장으로 남기기 어려운",
            "evidence": "서술형 답안의 식과 설명",
            "action": "식의 선택 이유와 결론을 짧은 문장으로 남기는 순서",
        },
        {
            "label": "시간 배분",
            "keywords": ("시간", "속도", "시험"),
            "persona": "아는 문제도 시험 시간 안에 끝내지 못하는",
            "evidence": "문항별 풀이 시간과 건너뛴 문제 기록",
            "action": "문항 난도에 따라 먼저 풀 문제와 다시 볼 문제를 나누는 순서",
        },
        {
            "label": "오답 재풀이",
            "keywords": ("오답", "다시", "반복"),
            "persona": "오답을 정리해도 같은 유형에서 다시 틀리는",
            "evidence": "첫 풀이와 힌트 뒤 재풀이 기록",
            "action": "힌트 없이 다시 푼 날짜와 성공 여부를 남기는 순서",
        },
    ),
    "영어": (
        {
            "label": "어휘 누적",
            "keywords": ("어휘", "단어", "암기"),
            "persona": "단어를 외워도 지문 안에서 뜻을 바로 고르기 어려운",
            "evidence": "단어 시험과 문맥 속 의미를 고친 기록",
            "action": "뜻·예문·지문 위치를 함께 확인하는 누적 복습 순서",
        },
        {
            "label": "문장 구조",
            "keywords": ("문법", "구문", "문장 구조"),
            "persona": "문법 개념은 알지만 긴 문장의 구조를 나누기 어려운",
            "evidence": "주어·동사·수식 관계를 표시한 문장",
            "action": "문장 뼈대를 표시하고 해석 근거를 설명하는 순서",
        },
        {
            "label": "독해 근거",
            "keywords": ("독해", "근거", "지문"),
            "persona": "지문은 읽지만 답의 근거를 문장으로 찾기 어려운",
            "evidence": "선택지와 근거 문장을 연결한 표시",
            "action": "질문 유형을 확인하고 근거 문장으로 돌아가는 순서",
        },
        {
            "label": "학교 본문",
            "keywords": ("본문", "교과서", "프린트"),
            "persona": "학교 본문을 외워도 변형 문장에서 흔들리는",
            "evidence": "교과서 본문과 학교 프린트의 변형 기록",
            "action": "본문 핵심 문장을 구조·어휘·변형 문장으로 확장하는 순서",
        },
        {
            "label": "서술형 교정",
            "keywords": ("서술", "영작", "쓰기"),
            "persona": "내용은 알지만 서술형 답안을 완전한 문장으로 쓰기 어려운",
            "evidence": "서술형 답안과 교정 전후 문장",
            "action": "핵심 표현을 고른 뒤 문장 구조와 철자를 다시 확인하는 순서",
        },
        {
            "label": "읽기 속도",
            "keywords": ("속도", "시간", "모의고사"),
            "persona": "문장을 꼼꼼히 읽다가 시험 후반의 시간이 부족한",
            "evidence": "지문별 소요 시간과 근거를 놓친 문항",
            "action": "질문을 먼저 확인하고 문단별 핵심을 짧게 남기는 순서",
        },
        {
            "label": "오답 분류",
            "keywords": ("오답", "틀린", "실수"),
            "persona": "틀린 이유를 어휘·구조·근거로 나누지 않고 답만 고치는",
            "evidence": "오답 옆에 남긴 이유와 다시 푼 결과",
            "action": "어휘·문장 구조·근거 오류를 구분해 다시 확인하는 순서",
        },
    ),
}


ELEMENTARY_SIGNALS = {
    "수학": (
        {
            "label": "연산 원리",
            "keywords": ("연산", "곱셈", "나눗셈", "계산"),
            "persona": "답은 구해도 계산 순서와 이유를 말로 설명하기 어려운",
            "evidence": "연산 과정과 계산 이유를 적은 풀이",
            "action": "계산 단계를 소리 내어 설명하고 같은 원리의 문항을 다시 푸는 순서",
        },
        {
            "label": "문장제 해석",
            "keywords": ("문장제", "조건", "문제 이해", "식 세우"),
            "persona": "계산은 가능하지만 문장제에서 필요한 조건을 고르기 어려운",
            "evidence": "문제의 수와 조건을 표시하고 식을 세운 흔적",
            "action": "묻는 것과 주어진 조건을 나눈 뒤 식의 이유를 설명하는 순서",
        },
        {
            "label": "분수·소수 개념",
            "keywords": ("분수", "소수", "수직선", "크기 비교"),
            "persona": "분수와 소수를 모양으로는 알지만 수의 크기와 뜻을 연결하기 어려운",
            "evidence": "그림·수직선·수로 같은 값을 나타낸 기록",
            "action": "그림과 수를 연결한 뒤 크기와 계산 이유를 말하는 순서",
        },
        {
            "label": "도형·측정",
            "keywords": ("도형", "각도", "길이", "넓이", "측정"),
            "persona": "도형의 이름은 알아도 성질과 측정 방법을 설명하기 어려운",
            "evidence": "도형에 조건과 측정 단위를 표시한 기록",
            "action": "도형의 성질과 단위를 확인하고 직접 그려 보는 순서",
        },
        {
            "label": "계산 정확도",
            "keywords": ("실수", "정확", "검산", "받아올림", "받아내림"),
            "persona": "풀이 방법은 알지만 자릿값과 부호를 자주 놓치는",
            "evidence": "계산 과정과 실수한 자리를 고친 기록",
            "action": "자릿값을 맞추고 역연산으로 검산하는 순서",
        },
        {
            "label": "단원 복습",
            "keywords": ("단원", "평가", "복습", "학교 진도"),
            "persona": "배운 날에는 풀지만 단원이 바뀌면 이전 내용을 잊기 쉬운",
            "evidence": "학교 진도표와 단원별 재풀이 날짜",
            "action": "학교 진도와 누적 복습 날짜를 나누어 배치하는 순서",
        },
        {
            "label": "오답 재풀이",
            "keywords": ("오답", "다시", "반복", "틀린"),
            "persona": "답을 고친 뒤 같은 유형에서 다시 막히는",
            "evidence": "첫 풀이와 힌트 뒤 재풀이 기록",
            "action": "힌트 없이 다시 푼 날짜와 성공 여부를 남기는 순서",
        },
    ),
    "영어": (
        {
            "label": "소리·철자 연결",
            "keywords": ("파닉스", "소리", "철자", "알파벳", "발음"),
            "persona": "알파벳은 알지만 소리와 철자를 바로 연결하기 어려운",
            "evidence": "낱말을 듣고 읽고 쓴 기록",
            "action": "소리를 듣고 철자를 찾은 뒤 낱말을 다시 읽는 순서",
        },
        {
            "label": "기초 어휘",
            "keywords": ("단어", "어휘", "낱말", "뜻"),
            "persona": "단어 뜻을 외워도 그림과 문장 속 의미를 고르기 어려운",
            "evidence": "낱말 카드와 문장 속 뜻을 확인한 기록",
            "action": "소리·뜻·그림·짧은 문장을 함께 연결하는 복습 순서",
        },
        {
            "label": "기본 문장",
            "keywords": ("문장", "표현", "어순", "문법"),
            "persona": "익힌 단어를 짧은 문장 순서로 배열하기 어려운",
            "evidence": "낱말을 배열하고 직접 말하거나 쓴 문장",
            "action": "문장 틀을 읽고 낱말을 바꾸어 말하고 쓰는 순서",
        },
        {
            "label": "듣기·말하기",
            "keywords": ("듣기", "말하기", "대화", "질문", "응답"),
            "persona": "배운 표현도 실제 질문을 들으면 바로 대답하기 어려운",
            "evidence": "질문을 듣고 짧게 응답한 기록",
            "action": "질문의 핵심어를 듣고 자신의 말로 짧게 답하는 순서",
        },
        {
            "label": "읽기 이해",
            "keywords": ("읽기", "본문", "이해", "내용"),
            "persona": "문장을 소리 내어 읽어도 핵심 내용을 설명하기 어려운",
            "evidence": "짧은 글을 읽고 그림이나 질문과 연결한 기록",
            "action": "문장별 뜻을 확인하고 글의 핵심을 한 문장으로 말하는 순서",
        },
        {
            "label": "쓰기·철자",
            "keywords": ("쓰기", "받아쓰기", "철자", "영작"),
            "persona": "말로는 아는 표현을 철자와 문장으로 옮기기 어려운",
            "evidence": "듣고 쓴 낱말과 고치기 전후 문장",
            "action": "소리를 나누어 철자를 쓰고 문장 첫 글자와 마침표를 확인하는 순서",
        },
        {
            "label": "짧은 반복 복습",
            "keywords": ("복습", "반복", "습관", "매일"),
            "persona": "수업에서는 기억하지만 며칠 뒤 배운 표현을 떠올리기 어려운",
            "evidence": "수업 뒤 다시 읽고 말한 날짜와 성공 기록",
            "action": "짧게 듣고 말하고 읽는 복습을 여러 날 나누어 실행하는 순서",
        },
    ),
}


def signals_for(config: CategoryConfig) -> tuple[dict, ...]:
    return (
        ELEMENTARY_SIGNALS[config.subject]
        if config.school_level == "초등"
        else SIGNALS[config.subject]
    )


STAGE_NOTES = {
    "고1수학학원": (
        "고교 첫 수학에서는 계산 절차뿐 아니라 사용한 개념의 이유를 설명하는 습관이 중요합니다.",
        "중학교에서 익힌 풀이가 통하더라도 새 정의와 기호를 정확히 읽는 과정부터 다시 확인해야 합니다.",
        "고1 공통 과목은 다음 단원의 바탕이 되므로 진도와 누적 복습 날짜를 함께 남기는 편이 안전합니다.",
        "첫 고교 내신은 문항 수보다 제한 시간 안에 풀이 근거를 유지하는지를 확인하는 기준이 됩니다.",
        "수행평가와 지필평가의 요구가 다를 수 있어 학교 자료를 실제로 본 뒤 설명 연습을 나누어야 합니다.",
        "무리한 선행보다 고1 과정의 개념어와 대표 문항을 스스로 설명하는 범위를 먼저 넓혀야 합니다.",
    ),
    "고1영어학원": (
        "고교 첫 영어에서는 단어 뜻 암기와 문장 안에서 의미를 고르는 능력을 따로 확인해야 합니다.",
        "중학교 때 짧은 문장에 익숙했다면 고1에서는 수식 관계와 접속 흐름을 표시하는 연습이 필요합니다.",
        "학교 본문은 암기한 문장을 변형해 물을 수 있으므로 구조와 어휘의 쓰임까지 설명해야 합니다.",
        "첫 고교 내신은 교과서·프린트·부교재의 비중을 확인한 뒤 자료별 복습 주기를 나누어야 합니다.",
        "서술형 준비에서는 외운 표현보다 질문에 맞는 근거를 골라 완전한 문장으로 쓰는지가 중요합니다.",
        "고1 어휘는 한 번의 시험으로 끝내지 않고 지문에서 다시 만난 날짜와 쓰임을 누적해야 합니다.",
    ),
    "고2수학학원": (
        "고2 수학에서는 현재 단원의 어려움과 고1 과정에서 남은 빈틈을 구분해야 계획이 흔들리지 않습니다.",
        "학교별 단원 순서가 다를 수 있으므로 학년명보다 실제 교재 목차와 최근 진도를 먼저 확인해야 합니다.",
        "누적 개념이 필요한 문항은 새 공식을 더 외우기보다 이전 정의를 어느 단계에서 쓰는지 설명해야 합니다.",
        "고2 내신은 문항 난도가 올라가므로 풀지 못한 문제와 시간 때문에 건너뛴 문제를 따로 기록해야 합니다.",
        "다음 시험까지 남은 기간에는 현재 단원 복습과 누적 빈틈 보완의 비율을 주마다 조정해야 합니다.",
        "진도를 따라가는 것과 혼자 다시 푸는 것은 다르므로 재풀이 결과를 다음 학습량의 기준으로 삼아야 합니다.",
    ),
    "고2영어학원": (
        "고2 영어에서는 지문 길이와 추상도가 높아져 단어·구조·근거 중 막힌 단계를 분리해야 합니다.",
        "학교별 본문과 부교재가 다를 수 있으므로 학년명보다 실제 범위표와 최근 시험지를 먼저 확인해야 합니다.",
        "누적 어휘는 뜻을 아는 수준과 문맥에서 정확한 의미를 고르는 수준을 나누어 기록해야 합니다.",
        "고2 내신 독해는 정답 선택보다 근거 문장을 찾아 오답 선택지와 비교하는 과정이 중요합니다.",
        "서술형과 변형 문장은 외운 문구만으로 대응하기 어려워 문장 구조를 바꾸어 쓰는 연습이 필요합니다.",
        "다음 시험 계획에는 새 지문 학습과 이전 오답 재확인의 시간을 각각 배치해야 합니다.",
    ),
    "중1수학학원": (
        "중1 수학에서는 계산 결과뿐 아니라 문자와 식이 뜻하는 관계를 설명하는 습관이 중요합니다.",
        "초등 과정의 계산이 익숙해도 중학교 개념어와 기호를 정확히 읽는 단계부터 확인해야 합니다.",
        "첫 중학교 시험은 단원별 대표 문항을 스스로 설명하고 제한 시간 안에 다시 푸는지 살펴보는 기준이 됩니다.",
        "학교 수업과 수행평가의 요구가 다를 수 있어 실제 교과서와 범위표를 본 뒤 연습 순서를 정해야 합니다.",
        "문제 수를 늘리기보다 개념 설명, 식 세우기, 계산 검토의 순서를 일정하게 만드는 편이 안전합니다.",
        "중1 과정은 다음 학년의 문자식·함수·도형 학습으로 이어지므로 오답의 원인과 재풀이 날짜를 남겨야 합니다.",
    ),
    "중1영어학원": (
        "중1 영어에서는 단어 뜻 암기와 문장 안에서 역할을 찾는 능력을 따로 확인해야 합니다.",
        "초등 영어의 짧은 표현에 익숙했다면 주어·동사와 수식 관계를 표시하는 연습이 필요합니다.",
        "중학교 첫 본문 평가는 암기한 문장을 바꾸어 물을 수 있어 구조와 어휘의 쓰임까지 설명해야 합니다.",
        "교과서·학교 프린트·수행평가의 비중을 확인한 뒤 자료별 복습 주기를 나누어야 합니다.",
        "서술형 준비에서는 외운 표현보다 질문에 맞는 단어와 문장 구조를 고르는 과정이 중요합니다.",
        "중1 어휘는 단어 시험에서 끝내지 않고 본문에서 다시 만난 날짜와 문맥을 함께 누적해야 합니다.",
    ),
    "중2수학학원": (
        "중2 수학에서는 현재 단원의 어려움과 중1 과정에서 남은 빈틈을 구분해야 계획이 흔들리지 않습니다.",
        "학교별 단원 순서가 다를 수 있으므로 학년명보다 실제 교재 목차와 최근 진도를 먼저 확인해야 합니다.",
        "문자식·함수·도형이 연결되는 문항은 공식을 더 외우기보다 어느 조건에서 쓰는지 설명해야 합니다.",
        "중2 내신은 풀이 단계가 길어지므로 풀지 못한 문제와 계산 때문에 틀린 문제를 따로 기록해야 합니다.",
        "다음 시험까지 남은 기간에는 현재 단원 복습과 누적 빈틈 보완의 비율을 주마다 조정해야 합니다.",
        "진도를 따라가는 것과 혼자 다시 푸는 것은 다르므로 재풀이 결과를 다음 학습량의 기준으로 삼아야 합니다.",
    ),
    "중2영어학원": (
        "중2 영어에서는 문장이 길어져 어휘·구조·본문 근거 중 막힌 단계를 분리해야 합니다.",
        "학교별 본문과 프린트가 다를 수 있으므로 학년명보다 실제 범위표와 최근 시험지를 먼저 확인해야 합니다.",
        "누적 어휘는 뜻을 아는 수준과 문맥에서 정확한 의미를 고르는 수준을 나누어 기록해야 합니다.",
        "중2 내신 독해는 정답 선택보다 근거 문장을 찾아 오답 선택지와 비교하는 과정이 중요합니다.",
        "서술형과 변형 문장은 외운 문구만으로 대응하기 어려워 문장 구조를 바꾸어 쓰는 연습이 필요합니다.",
        "다음 시험 계획에는 새 본문 학습과 이전 오답 재확인의 시간을 각각 배치해야 합니다.",
    ),
    "중3수학학원": (
        "중3 수학에서는 현재 내신 단원과 고교 과정으로 이어질 누적 개념을 함께 확인해야 합니다.",
        "학교 시험을 마무리하는 일정과 고교 수학을 준비하는 일정을 분리해 과도한 선행을 피해야 합니다.",
        "복합 문항은 공식을 외우는 데서 멈추지 않고 조건을 식과 도형으로 바꾸는 이유를 설명해야 합니다.",
        "중3 내신은 시간 배분과 서술형 풀이를 함께 요구하므로 건너뛴 문제와 설명이 끊긴 단계를 기록해야 합니다.",
        "고교 전환 전에는 새 단원을 많이 보는 것보다 중학교 핵심 개념을 혼자 다시 설명하는 범위를 넓혀야 합니다.",
        "시험 뒤 오답은 중학교 내신 보완과 고교 준비 항목으로 나누어 다음 학습 계획에 반영해야 합니다.",
    ),
    "중3영어학원": (
        "중3 영어에서는 중학교 본문 학습과 고교 지문을 읽을 기초 어휘·구조를 함께 확인해야 합니다.",
        "학교 시험 일정과 고교 영어 준비를 분리해 현재 내신 자료의 복습을 놓치지 않아야 합니다.",
        "누적 어휘는 뜻을 외운 개수보다 긴 문장 안에서 알맞은 의미를 고르는지를 기준으로 확인해야 합니다.",
        "중3 내신 독해는 근거 문장을 찾고 오답 선택지의 차이를 설명하는 과정까지 남겨야 합니다.",
        "고교 전환 전에는 문장 뼈대와 수식 관계를 표시하고 서술형 문장을 직접 고치는 연습이 필요합니다.",
        "다음 계획에는 중학교 오답 재확인과 고교 독해 준비 시간을 각각 배치해 우선순위를 분명히 해야 합니다.",
    ),
    "초3수학학원": (
        "초3 수학은 답만 맞히기보다 곱셈과 나눗셈의 계산 원리를 학생이 말로 설명하는 과정이 중요합니다.",
        "문장제에서는 숫자를 바로 계산하기 전에 묻는 것과 주어진 조건을 서로 다른 표시로 나누어 봅니다.",
        "분수는 모양을 외우는 데서 멈추지 않고 전체와 부분의 관계를 그림과 수로 함께 나타내야 합니다.",
        "길이·시간·들이를 다룰 때는 계산 결과와 단위를 짝지어 쓰는 습관을 확인해야 합니다.",
        "학교 단원평가 준비는 새 문제를 늘리기보다 틀린 이유를 말하고 며칠 뒤 다시 푸는 과정으로 이어져야 합니다.",
        "가정 복습은 긴 문제집 학습보다 대표 문항을 설명하고 계산을 검산하는 짧은 순서가 실행하기 쉽습니다.",
    ),
    "초3영어학원": (
        "초3 영어는 알파벳 이름을 아는 것과 낱말의 소리를 읽는 능력을 구분해 확인해야 합니다.",
        "처음 배우는 표현은 듣고 따라 말한 뒤 그림과 뜻을 연결하고 짧게 써 보는 순서가 도움이 됩니다.",
        "단어를 한 번에 많이 외우기보다 수업 뒤 여러 날 나누어 듣고 읽는 복습 기록을 남기는 편이 좋습니다.",
        "학교 영어 자료는 듣기·말하기 비중이 클 수 있어 교재 쪽수만으로 학습 상태를 판단하지 않습니다.",
        "틀린 철자는 정답만 베끼지 말고 들리는 소리와 빠진 글자를 나누어 다시 확인해야 합니다.",
        "영어를 처음 시작하는 학생에게는 정답률보다 소리 내어 읽고 짧게 대답하는 참여 과정을 먼저 봅니다.",
    ),
    "초4수학학원": (
        "초4 수학은 계산 단계가 길어지므로 답과 함께 중간 과정과 검산 흔적을 남기는 습관이 중요합니다.",
        "분수와 소수는 숫자 모양만 비교하지 말고 그림·수직선·생활 장면으로 크기의 뜻을 연결해야 합니다.",
        "여러 단계의 문장제는 한 번에 식을 세우기보다 먼저 구할 값과 마지막에 묻는 값을 나누어 봅니다.",
        "각도와 도형 단원에서는 눈대중보다 성질과 측정 단위를 말하고 직접 그려 보는 과정을 확인해야 합니다.",
        "학교 단원평가 뒤에는 계산 실수와 개념 오해, 조건 누락을 구분해 다음 복습 순서를 정합니다.",
        "문제 수를 늘리기 전에 대표 문항을 설명하고 조건이 바뀐 문항에도 같은 원리를 적용하는지 봅니다.",
    ),
    "초4영어학원": (
        "초4 영어는 익힌 단어를 짧은 문장 안에서 읽고 뜻을 설명하는 단계까지 확인해야 합니다.",
        "기본 문장은 통째로 외우기보다 낱말 하나를 바꾸어 말하고 쓰는 연습으로 구조를 익히는 편이 좋습니다.",
        "듣기에서는 모든 단어를 잡으려 하기보다 질문의 핵심어를 듣고 알맞게 응답하는지를 살펴봅니다.",
        "읽기와 쓰기는 따로 두지 않고 읽은 문장의 핵심 표현을 사용해 짧게 다시 쓰는 과정으로 연결합니다.",
        "학교 자료와 가정 복습의 표현이 다르면 공통 낱말과 문장 틀을 먼저 확인해 부담을 조절해야 합니다.",
        "한 번 외운 단어도 여러 문장에서 뜻을 고르고 철자를 다시 쓰는 날짜를 나누어 기록해야 합니다.",
    ),
}


DIVERSITY_NOTES = (
    "확인 결과는 맞음과 틀림만 적지 말고 학생이 혼자 설명한 범위까지 함께 남깁니다.",
    "한 번 이해한 내용은 이틀 뒤 같은 기준으로 다시 확인해야 실제 유지 여부를 알 수 있습니다.",
    "실행하지 못한 과제는 양을 더하기 전에 시작하지 못한 이유와 가능한 시간을 다시 정합니다.",
    "교재 진도와 학생 이해도가 다르면 진도표보다 설명과 재풀이 기록을 우선해 계획을 조정합니다.",
    "상담 질문은 추상적인 목표보다 이번 주에 확인할 자료와 행동을 중심으로 적는 편이 좋습니다.",
    "학생이 받은 힌트의 양도 기록해야 다음 재풀이에서 독립적으로 해결했는지 비교할 수 있습니다.",
    "복습 날짜를 미리 정해 두면 기억이 흐려진 뒤에도 같은 기준으로 다시 점검할 수 있습니다.",
    "학습량을 바꿀 때는 정답률과 함께 설명의 정확도와 완료 시간을 같이 비교해야 합니다.",
    "학교 일정이 바뀌면 기존 계획을 그대로 밀어붙이지 말고 확인 항목의 순서를 다시 배열합니다.",
    "학부모에게 전달할 기록은 진도, 막힌 이유, 다음 행동이 한눈에 구분되어야 합니다.",
    "학생이 질문하지 못했다면 이해한 것으로 단정하지 말고 짧은 설명이나 대표 문항으로 확인합니다.",
    "첫 주 계획은 작게 시작하되 확인 날짜와 수정 기준을 구체적으로 남겨야 합니다.",
    "반복 학습은 같은 문제를 보는 데서 끝내지 않고 조건이나 표현이 바뀌어도 적용하는지 확인합니다.",
)


DEEP_DIVERSITY_NOTES = (
    "최근 평가 자료에는 틀린 문항 옆에 막힌 단계를 짧게 적습니다. 다음 복습에서는 그 단계만 가린 채 학생이 스스로 이어 갈 수 있는지 확인합니다.",
    "학생 설명을 녹음하거나 한두 문장으로 받아 적으면 이해한 범위가 구체적으로 보입니다. 같은 내용을 다음 주에도 설명하는지 비교하면 유지 정도를 판단할 수 있습니다.",
    "교재 한 쪽을 끝내는 것보다 대표 문항 하나를 완전히 설명하는 것을 먼저 목표로 둡니다. 성공한 뒤에만 조건이나 표현이 다른 문항으로 범위를 넓힙니다.",
    "오답은 원인, 교정 행동, 재확인 날짜의 세 칸으로 나누어 기록합니다. 답만 고친 문항과 혼자 다시 해결한 문항을 구분해야 다음 과제량을 정하기 쉽습니다.",
    "학교 일정표에는 시험일뿐 아니라 수행평가와 과제 마감일도 표시합니다. 학습 계획은 빈 시간의 총량보다 실제로 시작할 수 있는 요일과 시간에 맞춥니다.",
    "상담에서 받은 설명은 학생이 집에서 다시 말해 보도록 합니다. 설명이 끊기는 지점이 있다면 새 진도보다 해당 개념이나 근거를 다음 질문으로 남깁니다.",
    "첫 진단과 한 달 뒤 기록은 같은 기준으로 비교해야 합니다. 자료와 질문이 달라지면 변화처럼 보일 수 있으므로 대표 문항과 설명 항목을 일정하게 유지합니다.",
    "수업 과제와 가정 복습의 역할을 구분합니다. 수업에서는 막힌 이유를 교정하고 집에서는 힌트 없이 다시 실행해 독립적으로 해낸 범위를 남깁니다.",
    "학습 계획을 바꿀 때는 학생에게 이유를 먼저 설명합니다. 무엇을 줄이고 무엇을 유지하는지 알면 과제의 우선순위를 이해하고 실행 결과도 정확히 보고할 수 있습니다.",
    "정답률이 비슷해도 풀이 시간과 설명의 정확도는 다를 수 있습니다. 세 지표를 함께 비교해 속도 문제인지 개념 문제인지 구분한 뒤 보완 순서를 정합니다.",
    "질문을 준비하지 못한 날에는 가장 오래 멈춘 문항을 표시해 옵니다. 상담에서는 그 흔적을 출발점으로 삼아 학생이 질문을 만드는 방법부터 연습할 수 있습니다.",
    "시험 직전에는 새 자료를 계속 추가하기보다 이미 틀린 항목의 재확인 비중을 높입니다. 다시 틀린 이유가 달라졌다면 오답 분류도 새로 정리합니다.",
    "학부모 피드백은 평가 문구보다 관찰 가능한 행동을 중심으로 요청합니다. 시작 시각, 완료한 범위, 혼자 설명한 내용과 다음 점검일이 있으면 가정에서도 이어 보기 쉽습니다.",
    "학생이 쉬운 문제만 반복하지 않도록 확인 문항의 난도를 세 단계로 나눕니다. 기본 확인 뒤 조건이 바뀐 문항과 설명이 필요한 문항을 순서대로 배치합니다.",
    "한 주 계획에는 보완 학습과 학교 진도를 별도 색으로 표시합니다. 두 계획이 겹치는 날에는 우선순위를 정해 미완료 항목이 계속 누적되지 않도록 조정합니다.",
    "자료가 많을수록 진단 기준을 단순하게 유지해야 합니다. 최근 시험지, 현재 교재, 학생 설명의 세 자료에서 공통으로 드러난 한 가지 문제부터 해결합니다.",
    "복습을 마쳤다는 말 대신 학생이 실제로 한 행동을 확인합니다. 다시 읽기, 설명하기, 힌트 없이 풀기 중 어떤 단계까지 했는지에 따라 다음 계획이 달라집니다.",
)


EXTRA_DIVERSITY_NOTES = (
    "다음 확인에서는 같은 자료를 사용해 학생의 설명이 더 짧고 정확해졌는지 비교합니다.",
    "막힌 지점이 달라졌다면 이전 계획을 반복하지 말고 새 원인에 맞춰 과제를 조정합니다.",
    "완료한 분량과 별개로 혼자 시작한 시각과 도움을 받은 지점을 함께 기록합니다.",
    "확인 문항은 배운 순서가 아니라 학생이 자주 놓치는 단계가 드러나도록 고릅니다.",
    "한 번에 여러 목표를 넣지 않고 이번 점검에서 바꿀 행동을 하나로 제한합니다.",
    "설명을 들은 직후의 정답과 며칠 뒤 혼자 푼 결과를 구분해 기록합니다.",
    "학생이 선택한 공부 순서도 물어보고 실제 계획과 다른 이유를 함께 정리합니다.",
    "틀린 문제를 지우기 전에 첫 생각과 수정한 근거를 나란히 남겨 변화 과정을 봅니다.",
    "학교 진도보다 보완이 늦어질 때는 필수 항목과 추가 항목을 구분해 부담을 조절합니다.",
    "상담에서 정한 용어를 가정에서도 동일하게 사용하면 학생이 피드백을 이해하기 쉽습니다.",
    "다음 수업 전까지 확인할 행동과 수업 중 교정할 행동을 서로 다른 칸에 적습니다.",
    "잘한 문항도 풀이 근거를 설명하게 해 우연한 정답과 실제 이해를 구분합니다.",
    "시험 뒤에는 점수 확인에서 멈추지 않고 다음 시험 전까지 남길 기록을 정합니다.",
    "계획을 줄이는 경우에도 반드시 유지할 핵심 복습 한 가지는 남겨 둡니다.",
    "학생이 질문한 표현을 그대로 기록하면 다음 상담에서 이해의 변화를 비교할 수 있습니다.",
)


def age_appropriate_note(note: str, config: CategoryConfig) -> str:
    """Keep shared planning notes natural for elementary-school readers."""
    if config.school_level != "초등":
        return note
    replacements = (
        (
            "학교 일정표에는 시험일뿐 아니라 수행평가와 과제 마감일도 표시합니다.",
            "학교 일정표에는 단원평가일뿐 아니라 준비물과 과제 마감일도 표시합니다.",
        ),
        (
            "시험 직전에는 새 자료를 계속 추가하기보다",
            "단원평가 전에는 새 학습지를 계속 추가하기보다",
        ),
        (
            "최근 시험지, 현재 교재, 학생 설명의 세 자료에서",
            "최근 단원평가나 학습지, 현재 교재, 학생 설명에서",
        ),
        (
            "시험 뒤에는 점수 확인에서 멈추지 않고 다음 시험 전까지",
            "단원평가 뒤에는 결과 확인에서 멈추지 않고 다음 단원평가 전까지",
        ),
        ("시험일", "단원평가일"),
        ("시험 주차", "단원평가 일정"),
    )
    result = note
    for before, after in replacements:
        result = result.replace(before, after)
    return result


class VisibleTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self.hidden_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() in {"script", "style"}:
            self.hidden_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style"} and self.hidden_depth:
            self.hidden_depth -= 1

    def handle_data(self, data: str) -> None:
        if not self.hidden_depth:
            self.parts.append(data)


def visible_text(source: str) -> str:
    parser = VisibleTextParser()
    parser.feed(source)
    return re.sub(r"\s+", " ", " ".join(parser.parts)).strip()


def stable_pick(key: str, slot: str, values: tuple[str, ...]) -> str:
    return shared.stable_pick(key, slot, list(values))


def stable_unique_picks(
    key: str,
    slot: str,
    values: tuple[str, ...],
    count: int,
) -> tuple[str, ...]:
    """Return deterministic candidates without reusing one on the same page."""
    if not 0 <= count <= len(values):
        raise ValueError(
            f"{slot}: requested {count} unique values from {len(values)} candidates"
        )
    ranked = sorted(
        values,
        key=lambda value: hashlib.sha256(
            f"{key}|{slot}|{value}".encode("utf-8")
        ).hexdigest(),
    )
    return tuple(ranked[:count])


def assert_unique_article_sentences(title: str, article: str) -> None:
    """Fail generation when a full sentence is repeated inside one article."""
    prose = re.sub(r"^## .+$", "", article, flags=re.MULTILINE)
    sentences = [
        re.sub(r"\s+", " ", match.group(0)).strip()
        for match in re.finditer(r"[^.!?\n]+[.!?]", prose)
    ]
    repeated = sorted(
        {sentence for sentence in sentences if sentences.count(sentence) > 1}
    )
    if repeated:
        raise ValueError(f"{title}: repeated article sentences={repeated[:3]}")


def final_jongseong(value: str) -> int:
    for character in reversed(value.strip()):
        code = ord(character)
        if 0xAC00 <= code <= 0xD7A3:
            return (code - 0xAC00) % 28
    return 0


def with_object(value: str) -> str:
    return value + ("을" if final_jongseong(value) else "를")


def with_conjunction(value: str) -> str:
    return value + ("과" if final_jongseong(value) else "와")


def with_direction(value: str) -> str:
    jongseong = final_jongseong(value)
    return value + ("로" if jongseong in {0, 8} else "으로")


def category_profiles() -> dict[str, dict[str, object]]:
    profiles: dict[str, dict[str, object]] = {}
    for config in CONFIGS:
        school_name = {
            "초등": "초등학교",
            "중등": "중학교",
            "고등": "고등학교",
        }[config.school_level]
        nationwide_links = (
            (
                ("고등수학학원", "수학 학습관리"),
                ("고등영어학원", "영어 학습관리"),
                ("고등영수학원", "영어·수학 학습관리"),
            )
            if config.school_level == "고등"
            else ()
        )
        profiles[config.category] = {
            "card_label": config.card_label,
            "card_description": config.card_description,
            "detail_eyebrow": config.detail_eyebrow,
            "category_eyebrow": config.category_eyebrow,
            "category_description": config.category_description,
            "category_short": f"{config.grade} {config.subject}",
            "badges": (
                config.grade,
                config.subject,
                "학교 단원 학습" if config.school_level == "초등" else "학교 내신",
                "오답 재학습",
            ),
            "aside_title": config.aside_title,
            "aside_copy": (
                "{local} 학생의 최근 시험지·교재·학교 일정과 설명 기록을 확인해 "
                "진단 뒤 바로 실행할 학습 순서를 정리합니다."
            ),
            "map_caption": (
                "{region} {district} {local}에서 "
                f"{config.grade} {config.subject}학원 상담을 준비할 때 실제 센터 "
                "위치와 평일 등원 동선을 함께 확인해 주세요."
            ),
            "related_description": (
                f"같은 동네의 학원 유형과 {school_name} 과목 안내, 같은 시군구와 광역권의 "
                f"{config.category} 페이지를 함께 정리했습니다."
            ),
            "topic_names": (
                f"{config.grade} {config.subject} 학습",
                f"{config.subject} 학교 내신",
                "현재 수준 진단",
                "오답 재학습",
                "상담 체크리스트",
            ),
            "parent_aside_title": config.parent_title,
            "parent_aside_copy": config.parent_copy,
            "llms_description": (
                "371개 동네별 원고 신호, 센터 주소, 해당 과목 가능 학년, "
                f"{school_name} 참고, 교습비와 지도 정보를 포함합니다."
            ),
            "faq_count": 5,
            "fixed_subject": config.subject,
            "fixed_grade": config.grade,
            "preserve_source_copy": True,
            "nationwide_links": nationwide_links,
        }
    return profiles


def load_sources(config: CategoryConfig) -> list[str]:
    path = SOURCE_DIR / config.source_name
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    values = [
        str(row[0].value or "").strip()
        for row in worksheet.iter_rows(min_col=1, max_col=1)
        if str(row[0].value or "").strip()
    ]
    if len(values) != 371:
        raise ValueError(f"{path.name}: expected 371 rows, found {len(values)}")
    return values


def ranked_signals(raw: str, config: CategoryConfig, key: str) -> list[dict]:
    text = visible_text(raw).lower()
    candidates = []
    for signal in signals_for(config):
        score = sum(text.count(keyword.lower()) for keyword in signal["keywords"])
        tie = hashlib.sha256(
            f"{key}|{signal['label']}".encode("utf-8")
        ).hexdigest()
        candidates.append((score, tie, signal))
    candidates.sort(key=lambda item: (-item[0], item[1]))
    return [item[2] for item in candidates]


def meta_description(
    config: CategoryConfig,
    local: str,
    primary: dict,
    secondary: dict,
    keyword_plan: keyword_strategy.KeywordPlan,
) -> str:
    candidates = keyword_strategy.meta_candidates(
        keyword_plan,
        local=local,
        primary_label=primary["label"],
        secondary_label=secondary["label"],
        key=f"{local}|{config.category}|meta",
        school_level=config.school_level,
    )
    for candidate in candidates:
        if 70 <= len(candidate) <= 100:
            return candidate
    raise ValueError(
        f"{local} {config.category}: no 70-100 character meta "
        f"lengths={[len(item) for item in candidates]}"
    )


def make_sections(
    raw: str,
    row: dict[str, str],
    config: CategoryConfig,
    index: int,
) -> dict[str, str]:
    local = shared.compact_text(row.get("근처 수업가능 동네"))
    district = shared.compact_text(row.get("시or구"))
    title = f"{local} {config.category}"
    key = f"{title}|{hashlib.sha256(raw.encode('utf-8')).hexdigest()}"
    primary, secondary, tertiary = ranked_signals(raw, config, key)[:3]
    keyword_plan = keyword_strategy.make_plan(
        category=config.category,
        grade=config.grade,
        subject=config.subject,
        local=local,
        primary_label=primary["label"],
        key=key,
        school_level=config.school_level,
    )
    keyword_strategy.validate_plan(
        keyword_plan,
        local=local,
        category=config.category,
    )
    grade_map = shared.grades_for(row)
    supported = config.grade in grade_map.get(config.subject, [])
    school_name = {
        "초등": "초등학교",
        "중등": "중학교",
        "고등": "고등학교",
    }[config.school_level]
    target_schools = shared.schools_for(row).get(config.school_level, [])
    school_text = "·".join(target_schools[:3])
    school_materials = (
        "교과서·학습지·단원 안내"
        if config.school_level == "초등"
        else "교과서·프린트·시험 범위표"
    )
    school_basis = (
        f"확인된 {school_name} 참고 범위는 {school_text}입니다. 학교명만으로 범위를 "
        f"단정하지 말고 실제 {school_materials}를 상담에서 다시 대조해야 합니다."
        if school_text
        else f"확인된 {school_name} 정보가 비어 있으므로 학교명을 추정하지 않습니다. "
        f"자녀 학교의 최근 {school_materials}를 준비해 상담 범위를 확인해야 합니다."
    )
    elementary = config.school_level == "초등"
    schedule_context = (
        f"{district} 생활권에서는 하교 시각, 교과서 단원, 단원평가와 방과 후 이동 시간을 함께 보아야 합니다. "
        if elementary
        else f"{district} 생활권에서는 하교 시각, 수행평가, 지필평가 주차와 이동 시간을 함께 보아야 합니다. "
    )
    school_heading = (
        f"교과서 단원과 {local} 주간 시간표를 맞추는 방법"
        if elementary
        else f"학교별 내신 자료와 {local} 주간 시간표를 맞추는 방법"
    )
    recent_learning_materials = (
        "최근 단원평가나 학습지, 현재 교재, 학교 단원 안내, 오답 기록과 주간 시간표"
        if elementary
        else "최근 시험지, 현재 교재, 학교 프린트, 오답 기록과 주간 시간표"
    )
    checklist_materials = (
        "① 최근 단원평가나 학습지와 교재 ② 학교 단원 안내와 주간 계획"
        if elementary
        else "① 최근 시험지와 교재 ② 학교 범위표와 프린트"
    )
    school_content_label = "학교 단원 학습 자료" if elementary else "학교 내신 자료"
    availability_intro = stable_pick(
        key,
        "availability-intro",
        (
            (
                f"센터에서 제공한 학년·과목 범위에는 {config.grade} {config.subject} 과정이 들어 있습니다. "
                "현재 모집 중인 반과 시간표는 상담 시점에 다시 확인해야 합니다."
            ),
            (
                f"제공된 센터 자료에서 {config.grade} {config.subject} 과정은 수업 가능 범위로 확인됩니다. "
                "실제 반 편성과 수업 시간은 상담할 때 최신 정보와 대조하세요."
            ),
            (
                f"현재 확보한 센터 정보에는 {config.grade} {config.subject} 과정이 가능 학년·과목으로 표시돼 있습니다. "
                "등록 전에는 운영 중인 반과 시간을 별도로 확인해야 합니다."
            ),
        )
        if supported
        else (
            (
                f"센터 자료에는 {config.grade} {config.subject} 가능 여부가 표시되지 않았습니다. "
                "제공 여부를 추정하지 말고 상담에서 대상 학년과 과목부터 확인하세요."
            ),
            (
                f"확보한 센터 정보만으로는 {config.grade} {config.subject} 과정 제공을 확인할 수 없습니다. "
                "가능하다고 전제하지 말고 상담할 때 학년·과목 범위를 먼저 물어보세요."
            ),
            (
                f"현재 자료에서는 {config.grade} {config.subject} 수업 가능 여부가 확인되지 않습니다. "
                "등록을 검토하기 전에 센터에 대상 학년과 운영 과목을 직접 확인해야 합니다."
            ),
        ),
    )
    availability_body = stable_pick(
        key,
        "availability-body",
        (
            (
                f"{config.grade} {config.subject} 과정이 센터의 확인된 가능 학년·과목에 포함된 상태입니다. "
                "반별 모집 여부와 실제 시간표는 방문 전에 문의하세요."
            ),
            (
                f"센터 자료에 {config.grade} {config.subject} 과정이 가능 범위로 기록돼 있지만 현재 반 개설을 뜻하지는 않습니다. "
                "상담에서 운영 반과 시작 가능한 시점을 확인하세요."
            ),
            (
                f"{config.grade} {config.subject} 과정은 제공 자료의 대상 범위에 해당합니다. "
                "반 구성과 수업 시각은 센터의 최신 안내를 기준으로 판단해야 합니다."
            ),
        )
        if supported
        else (
            (
                f"센터 자료에서 {config.grade} {config.subject} 과정의 제공 여부를 확인할 수 없습니다. "
                "상담 전에 대상 학년·과목을 먼저 문의하고 가능하다는 전제 없이 비교하세요."
            ),
            (
                f"{config.grade} {config.subject} 과정은 확보된 센터 정보의 가능 범위에 나타나지 않습니다. "
                "운영 여부와 등록 가능한 시점을 센터에 직접 물어봐야 합니다."
            ),
            (
                f"현재 정보에는 {config.grade} {config.subject} 과정 제공 여부가 비어 있습니다. "
                "학년과 과목이 모두 가능한지 확인한 뒤 상담 일정을 정하세요."
            ),
        ),
    )

    intro = stable_pick(
        key,
        "intro",
        (
            f"{local}에서 {primary['persona']} {config.grade} 학생이라면 문제 수보다 진단 순서가 먼저입니다. "
            f"결론부터 말하면 {primary['evidence']}부터 살핀 뒤 {with_direction(primary['action'])} 이어지는지 확인해야 합니다. {availability_intro}",
            f"{config.transition}에는 공부 시간을 늘리기 전에 막힌 지점을 구분해야 합니다. "
            f"{local} {config.grade} 학생의 {with_object(primary['evidence'])} 보고 {primary['label']}부터 보완하는 것이 첫 단계입니다. {availability_intro}",
            f"{local}에서 {config.grade} {with_object(config.subject)} 준비할 때는 점수 하나보다 학생이 풀이와 답의 근거를 설명하는 과정을 보아야 합니다. "
            f"{with_conjunction(primary['label'])} {with_object(secondary['label'])} 나누어 진단한 뒤 다음 행동을 정리하세요. {availability_intro}",
        ),
    )

    headings = (
        stable_pick(
            key,
            "heading-1",
            (
                f"{local} {config.grade} {config.subject}, 지금 막힌 지점부터 어떻게 찾을까요?",
                f"{primary['label']}에서 시작하는 {local} {config.grade} 진단 기준",
                f"최근 시험지로 찾는 {config.grade} {config.subject}의 첫 보완 지점",
            ),
        ),
        stable_pick(
            key,
            "heading-2",
            (
                f"{with_object(primary['label'])} 점수보다 설명 기록으로 확인하는 방법",
                f"{primary['evidence']}에서 읽어야 할 진단 신호",
                f"첫 상담에서 {with_object(primary['label'])} 구분하는 질문",
            ),
        ),
        stable_pick(
            key,
            "heading-3",
            (
                f"{secondary['label']}·{with_object(tertiary['label'])} 연결하는 주간 복습 설계",
                f"수업 뒤 {with_object(secondary['label'])} 다시 확인하는 실행 순서",
                f"{primary['label']} 진단을 다음 과제와 오답 계획으로 잇는 법",
            ),
        ),
        stable_pick(
            key,
            "heading-4",
            (
                f"{district} 학교 일정과 {local} 등원 계획을 함께 보는 이유",
                f"지역·학년·추천학생 기준: {local} {config.grade} 학습 일정",
                school_heading,
            ),
        ),
        stable_pick(
            key,
            "heading-5",
            (
                "학부모가 확인할 피드백과 가정 복습 기록",
                f"{config.grade} 학생의 설명과 재학습 결과를 확인하는 피드백",
                "진도표보다 다음 행동이 보이는 학부모 점검 기준",
            ),
        ),
        f"{title} 상담 전 체크리스트",
    )
    headings = keyword_strategy.apply_heading_plan(
        headings,
        keyword_plan,
        category=config.category,
        primary_label=primary["label"],
        secondary_label=secondary["label"],
        tertiary_label=tertiary["label"],
        key=key,
        school_level=config.school_level,
    )

    base_paragraphs = (
        (
            f"{primary['persona']} 학생에게는 새 진도를 서두르기보다 {with_object(primary['evidence'])} 먼저 확인하는 편이 안전합니다. "
            f"오류가 시작된 단계와 학생이 혼자 설명할 수 있는 범위를 나누면 {primary['label']}의 보완 순서가 선명해집니다.",
            f"{config.grade} 시기는 {config.transition}입니다. 최근 자료를 기준으로 현재 가능한 행동을 하나 정하고, "
            "수업 뒤 같은 기준으로 다시 확인해야 계획이 실제 학습으로 이어집니다.",
        ),
        (
            f"상담에서는 {with_object(primary['evidence'])} 펼쳐 놓고 정답 여부, 설명 과정, 다시 푼 결과를 따로 봅니다. "
            f"그 결과를 바탕으로 {with_object(primary['action'])} 첫 주의 확인 항목으로 정할 수 있습니다.",
            f"{secondary['label']}도 함께 보되 두 약점을 한꺼번에 고치려 하지 않습니다. "
            f"학생이 혼자 해낼 수 있는 단계부터 기록하고 다음 확인 날짜를 남기는 것이 {config.subject} 진단의 핵심입니다.",
        ),
        (
            f"주간 계획에는 수업 날짜만 적지 말고 {secondary['label']} 점검 순서와 {tertiary['label']} 재확인 순서를 각각 넣어야 합니다. "
            "한 번 설명한 내용이 과제와 재확인 문항에서도 유지되는지 확인하면 복습의 우선순위를 조정하기 쉽습니다.",
            f"과제량은 정답 수보다 학생이 혼자 다시 해낸 기록으로 조정합니다. {local} 학생의 학교 일정과 가정 복습 시간을 함께 적고, "
            "실행하지 못한 날에는 양을 늘리기보다 이유와 다음 행동을 먼저 정리합니다.",
        ),
        (
            school_basis,
            schedule_context
            +
            f"{local}에서 현실적으로 지킬 수 있는 복습 시간을 계산하고 {primary['label']} 확인 날짜를 학교 일정 앞에 배치하세요. {availability_body}",
        ),
        (
            f"학부모 피드백에는 진도와 숙제 여부만이 아니라 학생이 {with_object(primary['label'])} 어떻게 설명했는지, "
            "힌트 없이 다시 해낸 항목은 무엇인지, 다음 확인일은 언제인지가 포함되어야 합니다.",
            f"가정에서는 답을 알려 주기보다 학생이 {with_object(secondary['evidence'])} 다시 설명하도록 질문해 보세요. "
            "수업 기록과 가정의 관찰이 같은 기준으로 이어지면 계획을 바꿀 시점을 판단하기 쉽습니다.",
        ),
        (
            f"{checklist_materials} ③ {primary['evidence']} "
            "④ 평일 등원·복습 가능 시간 ⑤ 학생이 원하는 보완 목표를 준비합니다. 자료가 없으면 추정하지 말고 상담에서 확인할 질문부터 적습니다.",
            f"상담 뒤에는 {primary['label']}의 진단 근거, 첫 주 실행 과제, 다시 확인할 날짜, 학부모가 받을 피드백 형식을 메모하세요. "
            "등록 여부는 설명이 구체적인지와 학생 일정에 실행 가능한지를 함께 비교해 결정합니다.",
        ),
    )
    section_count = len(base_paragraphs)
    diversity_notes = stable_unique_picks(
        key, "section-diversity", DIVERSITY_NOTES, section_count
    )
    deep_diversity_notes = stable_unique_picks(
        key, "section-deep-diversity", DEEP_DIVERSITY_NOTES, section_count
    )
    extra_seed = int(
        hashlib.sha256(f"{key}|article-extra-count".encode("utf-8")).hexdigest()[:8],
        16,
    )
    extra_count = (
        1 + extra_seed % section_count
        if config.category == "중2영어학원"
        else extra_seed % (section_count + 1)
    )
    extra_sections = stable_unique_picks(
        key,
        "section-extra-placement",
        tuple(str(section_index) for section_index in range(section_count)),
        extra_count,
    )
    extra_notes = stable_unique_picks(
        key, "section-extra-diversity", EXTRA_DIVERSITY_NOTES, extra_count
    )
    middle2_english_bonus = ""
    if config.category == "중2영어학원" and extra_count == section_count:
        remaining_notes = tuple(
            note for note in EXTRA_DIVERSITY_NOTES if note not in extra_notes
        )
        middle2_english_bonus = stable_pick(
            key,
            "middle2-english-length-bonus",
            remaining_notes,
        )
    extra_by_section = {
        int(section_index): note
        for section_index, note in zip(extra_sections, extra_notes)
    }
    paragraph_rows = []
    for section_index, (first, second) in enumerate(base_paragraphs):
        expansion = [
            second,
            STAGE_NOTES[config.category][section_index],
            age_appropriate_note(diversity_notes[section_index], config),
            age_appropriate_note(deep_diversity_notes[section_index], config),
        ]
        if section_index in extra_by_section:
            expansion.append(
                age_appropriate_note(extra_by_section[section_index], config)
            )
        if middle2_english_bonus and section_index == section_count - 1:
            expansion.append(middle2_english_bonus)
        paragraph_rows.append((first, " ".join(expansion)))
    paragraphs = tuple(paragraph_rows)
    body_parts = [intro]
    for heading, section_paragraphs in zip(headings, paragraphs):
        body_parts.append(
            "## " + heading + "\n\n" + "\n\n".join(section_paragraphs)
        )
    article_body = "\n\n".join(body_parts)
    assert_unique_article_sentences(title, article_body)

    faqs = (
        (
            f"{local} {config.category} 상담에서는 무엇부터 확인하나요?",
            f"먼저 {with_object(primary['evidence'])} 확인합니다. {local} 학생이 {primary['label']}에서 막히는지 설명과 재확인 기록으로 구분한 뒤 첫 주 학습 순서를 정합니다.",
        ),
        (
            f"{local}에서 {config.grade} {config.subject} 내신 대비는 어떻게 시작하나요?",
            f"{title} 내신 대비는 최근 학교 범위표와 시험지를 기준으로 시작합니다. 공통 개념과 학교별 자료를 나누고 {secondary['label']}·{tertiary['label']}의 보완 순서를 시험 주차에 맞춰 정합니다.",
        ),
        (
            f"{title} 상담 전에 어떤 자료를 준비하면 좋나요?",
            f"{title} 상담에는 {recent_learning_materials}를 준비하세요. 자료가 부족하면 {with_object(primary['label'])} 확인할 대표 문항과 학생의 설명부터 살펴볼 수 있습니다.",
        ),
        (
            f"{local}의 {config.grade} {config.subject} 수업 가능 여부는 확인됐나요?",
            (
                f"센터가 제공한 대상 범위에는 {config.grade} {config.subject} 과정이 들어 있습니다. 현재 모집 중인 반과 시간표는 {local} 상담에서 다시 확인해 주세요."
                if supported
                else f"센터 자료에서 {config.grade} {config.subject} 과정의 운영 여부를 찾을 수 없습니다. 제공한다고 전제하지 말고 {local} 상담에서 대상 학년과 과목을 먼저 물어봐야 합니다."
            ),
        ),
        (
            f"{title} 학부모는 어떤 피드백을 요청하면 좋나요?",
            f"{title} 피드백은 진도보다 진단 근거와 다음 행동을 보여 주어야 합니다. {primary['label']}의 설명 기록, 혼자 다시 해낸 결과, 다음 점검 날짜가 함께 전달되는지 확인하는 것이 좋습니다.",
        ),
    )
    keyword_faq = keyword_strategy.answer_faq(
        keyword_plan,
        local=local,
        grade=config.grade,
        subject=config.subject,
        primary_evidence=primary["evidence"],
        secondary_label=secondary["label"],
        tertiary_label=tertiary["label"],
        key=key,
        school_level=config.school_level,
    )
    faqs = (faqs[0], keyword_faq, *faqs[2:])
    faq_text = "\n\n".join(
        f"Q{number}. {question}\nA{number}. {answer}"
        for number, (question, answer) in enumerate(faqs, 1)
    )
    consultation = (
        f"{local}의 {config.grade} 학생이 {primary['persona']} 상황을 가정했습니다. "
        f"상담에서 {with_object(primary['evidence'])} 함께 보고 첫 주에는 {primary['action']}만 실행한 뒤, "
        "다시 해낸 결과와 학교 일정을 기준으로 다음 계획을 조정하는지 확인하는 사례입니다."
    )
    source_status = (
        "최근 학습 자료와 확인된 센터 정보를 근거로 정리했습니다"
        if "<h1" in raw.lower()
        else "확인 가능한 센터 정보와 학년별 학습 기준을 바탕으로 정리했습니다"
    )
    summary = (
        f"{title} 안내는 {keyword_plan.secondary}, {primary['label']}, {secondary['label']}, {school_content_label}, "
        f"오답 재학습과 상담 체크리스트를 다룹니다. {source_status}"
    )
    return {
        "페이지타이틀": title,
        "메타설명": meta_description(
            config,
            local,
            primary,
            secondary,
            keyword_plan,
        ),
        "본문": article_body,
        "FAQ": faq_text,
        "학부모후기": (
            "※ 실제 이용 후기가 아닌 상담 준비 상황 예시입니다.\n- "
            + consultation
        ),
        "JSON-LD 요약": summary,
    }


def update_llms(config: CategoryConfig) -> None:
    path = shared.SITE / "llms.txt"
    source = path.read_text(encoding="utf-8")
    school_name = {
        "초등": "초등학교",
        "중등": "중학교",
        "고등": "고등학교",
    }[config.school_level]
    lines = (
        f"- {config.category} 지역 목록: {shared.absolute_url(shared.PARENT, config.category)}",
        f"- {config.category} 지역 페이지는 371개 동네별 원고 신호, 센터 주소, "
        f"{config.grade} {config.subject} 가능 여부, {school_name} 참고, 교습비와 지도 정보를 포함합니다.",
    )
    additions = [line for line in lines if line not in source]
    if additions:
        path.write_text(
            source.rstrip() + "\n" + "\n".join(additions) + "\n",
            encoding="utf-8",
            newline="\n",
        )


def configure_shared(config: CategoryConfig) -> None:
    shared.CATEGORY = config.category
    shared.PUBLISHED_DATE = (
        ELEMENTARY_PUBLISHED_AT
        if config.school_level == "초등"
        else EXISTING_PUBLISHED_AT
    )
    shared.UPDATED_AT = UPDATED_AT
    shared.ASSET_VERSION = "20260814-1"


def select_configs(
    category_names: list[str] | None = None,
    middle_only: bool = False,
    elementary_only: bool = False,
) -> tuple[CategoryConfig, ...]:
    if category_names:
        requested = set(category_names)
        configs = tuple(config for config in CONFIGS if config.category in requested)
    elif middle_only:
        configs = tuple(config for config in CONFIGS if config.school_level == "중등")
    elif elementary_only:
        configs = tuple(config for config in CONFIGS if config.school_level == "초등")
    else:
        configs = CONFIGS
    if not configs:
        raise ValueError("No categories selected")
    return configs


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate grade-specific subject academy pages."
    )
    scope = parser.add_mutually_exclusive_group()
    scope.add_argument(
        "--category",
        action="append",
        choices=tuple(config.category for config in CONFIGS),
        help="Generate one category; repeat the option to select more than one.",
    )
    scope.add_argument(
        "--middle-only",
        action="store_true",
        help="Generate only the six middle-school subject categories.",
    )
    scope.add_argument(
        "--elementary-only",
        action="store_true",
        help="Generate only the four elementary-school subject categories.",
    )
    return parser.parse_args()


def main(configs: tuple[CategoryConfig, ...] | None = None) -> None:
    active_configs = CONFIGS if configs is None else configs
    if not active_configs:
        raise ValueError("At least one category is required")
    shared.CATEGORY_CATALOG.update(category_profiles())
    rows = shared.read_csv(shared.COMMON / "센터정보 정리.csv")
    shared.enrich_center_rows(rows)
    if len(rows) != 371:
        raise ValueError(f"Expected 371 center rows, found {len(rows)}")

    source_sets = {
        config.category: load_sources(config) for config in active_configs
    }
    prepared: dict[str, list[dict[str, str]]] = {}
    fallback_counts: dict[str, int] = {}
    for config in active_configs:
        sources = source_sets[config.category]
        prepared[config.category] = [
            make_sections(raw, row, config, index)
            for index, (raw, row) in enumerate(zip(sources, rows))
        ]
        fallback_counts[config.category] = sum(
            "<h1" not in raw.lower() for raw in sources
        )
        target = shared.SITE / shared.PARENT / config.category
        target.mkdir(parents=True, exist_ok=True)

    representatives: dict[str, list[str]] = {}
    for config in active_configs:
        configure_shared(config)
        representatives[config.category] = shared.representative_paths(len(rows))

    # Two deterministic passes ensure every selected detail page receives
    # reciprocal sibling-category links after the category trees exist.
    for pass_number in (1, 2):
        for config in active_configs:
            configure_shared(config)
            target = shared.SITE / shared.PARENT / config.category
            for index, (sections, row) in enumerate(
                zip(prepared[config.category], rows)
            ):
                slug = shared.slug_ko(row["근처 수업가능 동네"])
                output = target / slug / "index.html"
                output.parent.mkdir(parents=True, exist_ok=True)
                rendered = shared.render_detail(
                    sections,
                    row,
                    representatives[config.category][index],
                    rows,
                )
                output.write_text(
                    re.sub(r"(?m)^[ \t]+$", "", rendered),
                    encoding="utf-8",
                    newline="\n",
                )
            (target / "index.html").write_text(
                shared.render_category_hub(rows),
                encoding="utf-8",
                newline="\n",
            )
            if pass_number == 2:
                update_llms(config)

    configure_shared(active_configs[-1])
    (shared.SITE / shared.PARENT / "index.html").write_text(
        shared.render_parent_hub(),
        encoding="utf-8",
        newline="\n",
    )
    generate_sitemap_robots.main()
    print(
        "generated "
        + " ".join(
            f"{config.category}=371(fallback={fallback_counts[config.category]})"
            for config in active_configs
        )
    )


if __name__ == "__main__":
    arguments = parse_args()
    main(
        select_configs(
            arguments.category,
            arguments.middle_only,
            arguments.elementary_only,
        )
    )
